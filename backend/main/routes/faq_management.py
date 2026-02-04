"""
FAQ管理路由 - 【核心文件】
~~~~~~~~~~

FAQ管理的API端点：
- 获取FAQ列表（分页、筛选、搜索）
- 获取单个FAQ详情
- 创建FAQ（支持AI关键词扩充）
- 更新FAQ
- 删除FAQ
- 批量操作（启用/禁用/删除）
- AI关键词建议
- 统计信息查询

核心接口：
- GET /api/faq/list: FAQ列表查询
- GET /api/faq/{id}: FAQ详情查询
- POST /api/faq: 创建FAQ（AI关键词自动生成）
- PUT /api/faq/{id}: 更新FAQ
- DELETE /api/faq/{id}: 删除FAQ
- POST /api/faq/batch: 批量操作
- POST /api/faq/keywords/suggest: AI关键词建议
- GET /api/faq/stats: 统计信息

AI功能：
- 创建FAQ时自动调用DeepSeek生成关键词
- 支持手动设置和AI自动生成关键词混合
- 关键词权重管理（手动=1.0，AI=0.5）

作者: 智糖团队
日期: 2025-01-21
"""

from flask import request, jsonify, Blueprint, send_file
from utils.jwt_helper import no_auth_required as token_required
from utils.logger import get_logger
from utils.database import get_db_connection  # 使用连接池
from services.knowledge_qa_service import get_knowledge_qa_service
from typing import Dict, Any, List, Optional
import json
import pymysql
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime

logger = get_logger(__name__)

# 创建Blueprint
faq_bp = Blueprint('faq_management', __name__, url_prefix='/api/faq')

# 获取知识问答服务实例
knowledge_service = get_knowledge_qa_service()


def safe_strip(value, default=''):
    """安全地处理字符串，如果为None则返回默认值"""
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip()
    return str(value).strip() if value else default


def extract_keywords_with_ai(question: str, answer: str, existing_keywords: List[str] = None) -> List[Dict[str, Any]]:
    """
    使用AI提取和扩充关键词

    Args:
        question: 问题文本
        answer: 答案文本
        existing_keywords: 已有的关键词列表

    Returns:
        List[Dict]: 关键词列表，包含keyword, type, weight
    """
    try:
        from services.deepseek_service import get_deepseek_service

        deepseek_service = get_deepseek_service()

        # 构建关键词提取提示词
        content = f"""
请为以下问答对提取关键词：

问题：{question}

答案：{answer[:500]}...  # 限制答案长度

要求：
1. 提取3-8个核心关键词
2. 关键词应该是名词或名词短语
3. 关键词应该与儿童青少年糖尿病管理相关
4. 每个关键词2-4个字为宜
5. 返回格式：用逗号分隔的关键词列表

示例格式：胰岛素,剂量计算,血糖控制,饮食管理
"""

        # 调用AI生成关键词
        messages = [
            {
                "role": "system",
                "content": "你是一个专业的医疗知识专家，擅长为医疗问答内容提取关键词。请返回简洁的关键词列表。"
            },
            {
                "role": "user",
                "content": content
            }
        ]

        response = deepseek_service.chat_completion(messages, stream=False, max_tokens=200)

        if response and 'choices' in response and len(response['choices']) > 0:
            ai_response = response['choices'][0]['message']['content'].strip()

            # 解析AI返回的关键词
            ai_keywords = []
            if ai_response:
                # 清理响应文本
                ai_response = ai_response.replace('关键词：', '').replace('关键字：', '').strip()
                ai_response = ai_response.replace('，', ',').replace('；', ',').replace('；', ',')

                # 按逗号分割
                keywords = [kw.strip() for kw in ai_response.split(',') if kw.strip()]

                for kw in keywords[:8]:  # 限制最多8个关键词
                    if 2 <= len(kw) <= 10 and kw not in (existing_keywords or []):
                        ai_keywords.append(kw)

            logger.info(f"🤖 AI提取关键词: {ai_keywords}")

            # 构建关键词对象
            keyword_objects = []

            # 首先添加现有关键词（如果有）
            if existing_keywords:
                for kw in existing_keywords:
                    keyword_objects.append({
                        'keyword': kw,
                        'type': 'manual',
                        'weight': 1.0
                    })

            # 添加AI提取的关键词
            for kw in ai_keywords:
                keyword_objects.append({
                    'keyword': kw,
                    'type': 'auto',
                    'weight': 0.5
                })

            return keyword_objects

        else:
            logger.warning("❌ AI关键词提取失败，返回空结果")
            return []

    except Exception as e:
        logger.error(f"❌ AI关键词提取异常: {str(e)}")
        return []


@faq_bp.route('/list', methods=['GET'], endpoint='get_faq_list')
@token_required
def get_faq_list(user_id):
    """
    获取FAQ列表

    Headers:
        Authorization: Bearer <token>

    Query Parameters:
        page: 页码 (默认1)
        page_size: 每页数量 (默认20)
        category: 分类筛选
        status: 状态筛选 (1=启用, 0=禁用)
        search: 搜索关键词
        source: 来源筛选

    Returns:
        JSON: FAQ列表及分页信息
    """
    try:
        # 获取查询参数，安全处理空字符串
        def safe_int(value, default=0):
            if not value or not str(value).strip():
                return default
            try:
                return int(value)
            except (ValueError, TypeError):
                return default
        
        page = safe_int(request.args.get('page'), default=1)
        page_size = safe_int(request.args.get('page_size'), default=20)
        category = request.args.get('category')
        status = request.args.get('status')
        search = request.args.get('search')
        source = request.args.get('source')

        offset = (page - 1) * page_size

        # 构建查询条件
        where_clauses = []
        params = []

        if category:
            where_clauses.append("category = %s")
            params.append(category)

        if status is not None:
            where_clauses.append("status = %s")
            params.append(int(status))

        if source:
            where_clauses.append("source = %s")
            params.append(source)

        if search:
            where_clauses.append("(question LIKE %s OR answer LIKE %s)")
            search_param = f"%{search}%"
            params.extend([search_param, search_param])

        where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 查询总数
            count_sql = f"SELECT COUNT(*) as total FROM faq_list{where_sql}"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']

            # 查询数据
            sql = f"""
                SELECT
                    f.id,
                    f.question,
                    f.answer,
                    f.category,
                    f.source,
                    f.status,
                    f.sort_order,
                    f.view_count,
                    f.like_count,
                    f.is_manual,
                    f.description,
                    f.created_at,
                    f.updated_at,
                    GROUP_CONCAT(
                        CONCAT(k.keyword, ':', k.keyword_type, ':', k.weight)
                        ORDER BY k.weight DESC, k.keyword
                    ) as keywords_str
                FROM faq_list f
                LEFT JOIN faq_list_keys k ON f.id = k.faq_id
                {where_sql}
                GROUP BY f.id
                ORDER BY f.sort_order ASC, f.created_at DESC, f.id DESC
                LIMIT %s OFFSET %s
            """
            params.extend([page_size, offset])
            cursor.execute(sql, params)
            records = cursor.fetchall()

            # 处理关键词
            for record in records:
                if record['keywords_str']:
                    keywords = []
                    for kw_str in record['keywords_str'].split(','):
                        if ':' in kw_str:
                            keyword, kw_type, weight = kw_str.split(':', 2)
                            keywords.append({
                                'keyword': keyword,
                                'type': kw_type,
                                'weight': float(weight)
                            })
                    record['keywords'] = keywords
                else:
                    record['keywords'] = []

                # 移除临时字段
                del record['keywords_str']

        conn.close()

        return jsonify({
            'success': True,
            'data': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size,
                'items': records
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ 获取FAQ列表失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/<int:faq_id>', methods=['GET'], endpoint='get_faq_detail')
@token_required
def get_faq_detail(user_id, faq_id):
    """
    获取单个FAQ详情

    Headers:
        Authorization: Bearer <token>

    Path Parameters:
        faq_id: FAQ ID

    Returns:
        JSON: FAQ详情
    """
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 查询FAQ详情
            sql = """
                SELECT
                    f.id,
                    f.question,
                    f.answer,
                    f.category,
                    f.source,
                    f.status,
                    f.sort_order,
                    f.view_count,
                    f.like_count,
                    f.is_manual,
                    f.description,
                    f.created_at,
                    f.updated_at,
                    GROUP_CONCAT(
                        CONCAT(k.keyword, ':', k.keyword_type, ':', k.weight)
                        ORDER BY k.weight DESC, k.keyword
                    ) as keywords_str
                FROM faq_list f
                LEFT JOIN faq_list_keys k ON f.id = k.faq_id
                WHERE f.id = %s
                GROUP BY f.id
            """
            cursor.execute(sql, (faq_id,))
            record = cursor.fetchone()

            if not record:
                return jsonify({
                    'code': 404,
                    'data': {},
                    'success': False,
                    'message': 'FAQ不存在'
                }), 404

            # 处理关键词
            if record['keywords_str']:
                keywords = []
                for kw_str in record['keywords_str'].split(','):
                    if ':' in kw_str:
                        keyword, kw_type, weight = kw_str.split(':', 2)
                        keywords.append({
                            'keyword': keyword,
                            'type': kw_type,
                            'weight': float(weight)
                        })
                record['keywords'] = keywords
            else:
                record['keywords'] = []

            # 移除临时字段
            del record['keywords_str']

        conn.close()

        return jsonify({
            'success': True,
            'data': record
        }), 200

    except Exception as e:
        logger.error(f"❌ 获取FAQ详情失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('', methods=['POST'], endpoint='create_faq')
@token_required
def create_faq(user_id):
    """
    创建FAQ

    Headers:
        Authorization: Bearer <token>

    Body:
        {
            "question": "问题内容",
            "answer": "答案内容",
            "category": "分类（可选）",
            "source": "来源（可选）",
            "keywords": ["关键词1", "关键词2"]（可选）,
            "status": 1（可选，1=启用，0=禁用）,
            "sort_order": 0（可选）,
            "description": "描述（可选）",
            "use_ai_keywords": true（可选，是否使用AI扩充关键词）
        }

    Returns:
        JSON: 创建结果
    """
    try:
        data = request.get_json()

        # 必填字段验证（安全处理 None 值）
        question = safe_strip(data.get('question'))
        answer = safe_strip(data.get('answer'))

        if not question:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '问题内容不能为空'
            }), 400

        if not answer:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '答案内容不能为空'
            }), 400

        # 可选字段（安全处理 None 值）
        category = safe_strip(data.get('category'))
        source = safe_strip(data.get('source'))
        keywords = data.get('keywords', [])
        status = int(data.get('status', 1))
        sort_order = int(data.get('sort_order', 0))
        description = safe_strip(data.get('description'))
        use_ai_keywords = data.get('use_ai_keywords', True)

        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 检查是否已存在相同问题
            cursor.execute('SELECT id FROM faq_list WHERE question = %s', (question,))
            if cursor.fetchone():
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': '相同问题已存在'
                }), 400

            # 准备关键词
            keyword_objects = []

            # 添加用户提供的关键词
            for kw in keywords:
                if isinstance(kw, str) and kw.strip():
                    keyword_objects.append({
                        'keyword': kw.strip(),
                        'type': 'manual',
                        'weight': 1.0
                    })

            # 使用AI扩充关键词
            if use_ai_keywords:
                logger.info("🤖 开始AI关键词扩充...")
                ai_keywords = extract_keywords_with_ai(question, answer, [kw['keyword'] for kw in keyword_objects])
                keyword_objects.extend(ai_keywords)

            # 创建FAQ
            sql = """
                INSERT INTO faq_list
                (question, answer, category, source, status, sort_order, description, is_manual)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (
                question, answer, category or None, source or None,
                status, sort_order, description or None, True
            ))

            faq_id = cursor.lastrowid

            # 添加关键词
            if keyword_objects:
                for kw_obj in keyword_objects:
                    try:
                        cursor.execute(
                            'INSERT INTO faq_list_keys (faq_id, keyword, keyword_type, weight) VALUES (%s, %s, %s, %s)',
                            (faq_id, kw_obj['keyword'], kw_obj['type'], kw_obj['weight'])
                        )
                    except pymysql.IntegrityError:
                        # 关键词已存在，跳过
                        pass

            conn.commit()

        conn.close()

        # 重新加载知识库（可选，保持数据一致性）
        try:
            knowledge_service._load_knowledge_base()
            logger.info("✅ 知识库已重新加载")
        except Exception as e:
            logger.warning(f"⚠️ 知识库重新加载失败: {str(e)}")

        return jsonify({
            'success': True,
            'message': 'FAQ创建成功',
            'data': {
                'id': faq_id,
                'keywords_count': len(keyword_objects)
            }
        }), 201

    except Exception as e:
        logger.error(f"❌ 创建FAQ失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/', methods=['PUT'], endpoint='update_faq_by_body')
@faq_bp.route('', methods=['PUT'], endpoint='update_faq_by_body_alias')
@token_required
def update_faq_by_body(user_id):
    """
    更新FAQ（从请求体获取ID）

    Headers:
        Authorization: Bearer <token>

    Body:
        {
            "id": 1,
            "question": "新问题内容",
            "answer": "新答案内容",
            "keywords": [{"keyword": "...", "type": "manual", "weight": 1.0}],
            ...
        }

    Returns:
        JSON: 更新结果
    """
    try:
        data = request.get_json()
        
        # 从请求体获取ID
        faq_id = data.get('id')
        if not faq_id:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '缺少FAQ ID'
            }), 400
        
        # 调用原有的更新函数
        return update_faq(user_id, faq_id)
        
    except Exception as e:
        logger.error(f"❌ 更新FAQ失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/<int:faq_id>', methods=['PUT'], endpoint='update_faq')
@token_required
def update_faq(user_id, faq_id):
    """
    更新FAQ

    Headers:
        Authorization: Bearer <token>

    Path Parameters:
        faq_id: FAQ ID

    Body:
        {
            "question": "新问题内容",
            "answer": "新答案内容",
            "category": "新分类",
            "source": "新来源",
            "keywords": ["新关键词1", "新关键词2"],
            "status": 1,
            "sort_order": 0,
            "description": "新描述"
        }

    Returns:
        JSON: 更新结果
    """
    try:
        data = request.get_json()

        # 验证必填字段
        question = data.get('question', '') or ''
        question = question.strip() if isinstance(question, str) else ''
        answer = data.get('answer', '') or ''
        answer = answer.strip() if isinstance(answer, str) else ''

        if not question:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '问题内容不能为空'
            }), 400

        if not answer:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '答案内容不能为空'
            }), 400

        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 检查FAQ是否存在
            cursor.execute('SELECT id FROM faq_list WHERE id = %s', (faq_id,))
            if not cursor.fetchone():
                return jsonify({
                    'code': 404,
                    'data': {},
                    'success': False,
                    'message': 'FAQ不存在'
                }), 404

            # 构建更新字段
            update_fields = []
            update_values = []

            # 必填字段
            update_fields.extend(['question = %s', 'answer = %s'])
            update_values.extend([question, answer])

            # 可选字段（安全处理 None 值）
            optional_fields = {
                'category': safe_strip(data.get('category')),
                'source': safe_strip(data.get('source')),
                'status': int(data.get('status', 1)),
                'sort_order': int(data.get('sort_order', 0)),
                'description': safe_strip(data.get('description'))
            }

            for field, value in optional_fields.items():
                if field in ['category', 'source', 'description'] and value == '':
                    value = None
                update_fields.append(f"{field} = %s")
                update_values.append(value)

            update_values.append(faq_id)

            # 更新FAQ
            sql = f"UPDATE faq_list SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(sql, update_values)

            # 处理关键词更新
            keywords = data.get('keywords', [])
            if keywords is not None:  # 允许空列表来清空关键词
                # 删除现有关键词
                cursor.execute('DELETE FROM faq_list_keys WHERE faq_id = %s', (faq_id,))

                # 添加新关键词
                for kw in keywords:
                    # 支持两种格式：字符串或对象
                    if isinstance(kw, str) and kw.strip():
                        keyword = kw.strip()
                        kw_type = 'manual'
                        weight = 1.0
                    elif isinstance(kw, dict):
                        keyword = kw.get('keyword', '').strip()
                        kw_type = kw.get('type', 'manual')
                        weight = float(kw.get('weight', 1.0))
                    else:
                        continue
                    
                    if keyword:
                        try:
                            cursor.execute(
                                'INSERT INTO faq_list_keys (faq_id, keyword, keyword_type, weight) VALUES (%s, %s, %s, %s)',
                                (faq_id, keyword, kw_type, weight)
                            )
                        except pymysql.IntegrityError:
                            pass

            conn.commit()

        conn.close()

        # 重新加载知识库
        try:
            knowledge_service._load_knowledge_base()
            logger.info("✅ 知识库已重新加载")
        except Exception as e:
            logger.warning(f"⚠️ 知识库重新加载失败: {str(e)}")

        return jsonify({
            'success': True,
            'message': 'FAQ更新成功'
        }), 200

    except Exception as e:
        logger.error(f"❌ 更新FAQ失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/<int:faq_id>', methods=['DELETE'], endpoint='delete_faq')
@token_required
def delete_faq(user_id, faq_id):
    """
    删除FAQ

    Headers:
        Authorization: Bearer <token>

    Path Parameters:
        faq_id: FAQ ID

    Returns:
        JSON: 删除结果
    """
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 检查FAQ是否存在
            cursor.execute('SELECT id FROM faq_list WHERE id = %s', (faq_id,))
            if not cursor.fetchone():
                return jsonify({
                    'code': 404,
                    'data': {},
                    'success': False,
                    'message': 'FAQ不存在'
                }), 404

            # 删除FAQ（级联删除关键词）
            cursor.execute('DELETE FROM faq_list WHERE id = %s', (faq_id,))
            conn.commit()

        conn.close()

        # 重新加载知识库
        try:
            knowledge_service._load_knowledge_base()
            logger.info("✅ 知识库已重新加载")
        except Exception as e:
            logger.warning(f"⚠️ 知识库重新加载失败: {str(e)}")

        return jsonify({
            'success': True,
            'message': 'FAQ删除成功'
        }), 200

    except Exception as e:
        logger.error(f"❌ 删除FAQ失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/batch', methods=['POST'], endpoint='batch_operation')
@token_required
def batch_operation(user_id):
    """
    批量操作FAQ

    Headers:
        Authorization: Bearer <token>

    Body:
        {
            "operation": "enable|disable|delete",
            "faq_ids": [1, 2, 3]
        }

    Returns:
        JSON: 批量操作结果
    """
    try:
        data = request.get_json()
        operation = data.get('operation')
        faq_ids = data.get('faq_ids', [])

        if not operation or not faq_ids:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '操作类型和FAQ ID列表不能为空'
            }), 400

        if operation not in ['enable', 'disable', 'delete']:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '无效的操作类型'
            }), 400

        conn = get_db_connection()
        with conn.cursor() as cursor:
            if operation == 'delete':
                # 批量删除
                placeholders = ','.join(['%s'] * len(faq_ids))
                cursor.execute(f'DELETE FROM faq_list WHERE id IN ({placeholders})', faq_ids)
            else:
                # 批量启用/禁用
                status = 1 if operation == 'enable' else 0
                placeholders = ','.join(['%s'] * len(faq_ids))
                cursor.execute(f'UPDATE faq_list SET status = %s WHERE id IN ({placeholders})', [status] + faq_ids)

            conn.commit()
            affected_rows = cursor.rowcount

        conn.close()

        # 重新加载知识库
        try:
            knowledge_service._load_knowledge_base()
            logger.info("✅ 知识库已重新加载")
        except Exception as e:
            logger.warning(f"⚠️ 知识库重新加载失败: {str(e)}")

        return jsonify({
            'success': True,
            'message': f'批量{operation}成功',
            'data': {
                'affected_count': affected_rows
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ 批量操作失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/keywords/suggest', methods=['POST'], endpoint='suggest_keywords')
@token_required
def suggest_keywords(user_id):
    """
    AI关键词建议

    Headers:
        Authorization: Bearer <token>

    Body:
        {
            "question": "问题内容",
            "answer": "答案内容",
            "existing_keywords": ["已有关键词"]（可选）
        }

    Returns:
        JSON: AI建议的关键词
    """
    try:
        data = request.get_json()

        question = data.get('question', '').strip()
        answer = data.get('answer', '').strip()
        existing_keywords = data.get('existing_keywords', [])

        if not question or not answer:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '问题和答案内容不能为空'
            }), 400

        # 调用AI关键词提取
        suggested_keywords = extract_keywords_with_ai(question, answer, existing_keywords)

        return jsonify({
            'success': True,
            'data': {
                'suggested_keywords': suggested_keywords,
                'count': len(suggested_keywords)
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ AI关键词建议失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/export', methods=['GET'], endpoint='export_faq')
@token_required
def export_faq(user_id):
    """
    导出FAQ为Excel文件

    Headers:
        Authorization: Bearer <token>

    Query Parameters:
        category: 分类筛选（可选）
        status: 状态筛选（可选）
        source: 来源筛选（可选）

    Returns:
        Excel文件下载
    """
    try:
        # 获取筛选参数
        category = request.args.get('category')
        status = request.args.get('status')
        source = request.args.get('source')

        # 构建查询条件
        where_clauses = []
        params = []

        if category and category.strip():
            where_clauses.append("f.category = %s")
            params.append(category)

        if status and status.strip():
            try:
                status_int = int(status)
                where_clauses.append("f.status = %s")
                params.append(status_int)
            except ValueError:
                # 忽略无效的status值
                pass

        if source and source.strip():
            where_clauses.append("f.source = %s")
            params.append(source)

        where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 查询FAQ数据
            sql = f"""
                SELECT
                    f.id,
                    f.question,
                    f.answer,
                    f.category,
                    f.source,
                    f.status,
                    f.sort_order,
                    GROUP_CONCAT(k.keyword ORDER BY k.keyword SEPARATOR ',') as keywords
                FROM faq_list f
                LEFT JOIN faq_list_keys k ON f.id = k.faq_id
                {where_sql}
                GROUP BY f.id
                ORDER BY f.sort_order ASC, f.created_at DESC, f.id DESC
            """
            cursor.execute(sql, params)
            records = cursor.fetchall()

        conn.close()

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "FAQ列表"

        # 设置表头
        headers = ['id', 'question', 'answer', 'category', 'source', 'keywords', 'status', 'sort_order']
        ws.append(headers)

        # 写入数据
        for record in records:
            ws.append([
                record['id'],
                record['question'],
                record['answer'],
                record['category'] or '',
                record['source'] or '',
                record['keywords'] or '',  # 关键词已经是逗号分隔的字符串
                record['status'],
                record['sort_order']
            ])

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'faq_export_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"❌ 导出FAQ失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/import', methods=['POST'], endpoint='import_faq')
@token_required
def import_faq(user_id):
    """
    从Excel文件导入FAQ

    Headers:
        Authorization: Bearer <token>

    Body:
        multipart/form-data
        file: Excel文件 (.xlsx)

    Excel格式要求:
        第一行为表头: question, answer, category, source, keywords, status, sort_order
        keywords列格式: 逗号分隔的关键词字符串

    Returns:
        JSON: 导入结果
    """
    try:
        # 检查文件是否存在
        if 'file' not in request.files:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '未找到上传文件'
            }), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '文件名为空'
            }), 400

        # 验证文件类型
        if not file.filename.endswith('.xlsx'):
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '仅支持.xlsx格式的Excel文件'
            }), 400

        # 读取Excel文件
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': f'Excel文件解析失败: {str(e)}'
            }), 400

        # 读取表头
        headers = [cell.value for cell in ws[1]]
        
        # 验证必需的列
        required_columns = ['question', 'answer']
        for col in required_columns:
            if col not in headers:
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': f'缺少必需列: {col}'
                }), 400

        # 获取列索引
        col_indices = {header: idx for idx, header in enumerate(headers)}

        # 导入统计
        total = 0
        success_count = 0
        failed_count = 0
        errors = []

        conn = get_db_connection()
        
        try:
            with conn.cursor() as cursor:
                # 遍历数据行（跳过表头）
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    total += 1
                    
                    try:
                        # 提取数据
                        question = row[col_indices['question']] if 'question' in col_indices else None
                        answer = row[col_indices['answer']] if 'answer' in col_indices else None
                        category = row[col_indices['category']] if 'category' in col_indices else None
                        source = row[col_indices['source']] if 'source' in col_indices else None
                        keywords_str = row[col_indices['keywords']] if 'keywords' in col_indices else None
                        status = row[col_indices['status']] if 'status' in col_indices else 1
                        sort_order = row[col_indices['sort_order']] if 'sort_order' in col_indices else 0

                        # 验证必填字段
                        if not question or not str(question).strip():
                            errors.append({'row': row_idx, 'reason': '问题内容为空'})
                            failed_count += 1
                            continue

                        if not answer or not str(answer).strip():
                            errors.append({'row': row_idx, 'reason': '答案内容为空'})
                            failed_count += 1
                            continue

                        question = str(question).strip()
                        answer = str(answer).strip()

                        # 检查是否已存在相同问题
                        cursor.execute('SELECT id FROM faq_list WHERE question = %s', (question,))
                        if cursor.fetchone():
                            errors.append({'row': row_idx, 'reason': '问题已存在'})
                            failed_count += 1
                            continue

                        # 处理可选字段
                        category = str(category).strip() if category else None
                        source = str(source).strip() if source else None
                        status = int(status) if status is not None else 1
                        sort_order = int(sort_order) if sort_order is not None else 0

                        # 插入FAQ
                        insert_sql = """
                            INSERT INTO faq_list
                            (question, answer, category, source, status, sort_order, is_manual)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(insert_sql, (
                            question, answer, category, source, status, sort_order, True
                        ))
                        faq_id = cursor.lastrowid

                        # 处理关键词
                        if keywords_str and str(keywords_str).strip():
                            keywords_str = str(keywords_str).strip()
                            # 解析逗号分隔的关键词
                            keywords = [kw.strip() for kw in keywords_str.split(',') if kw.strip()]
                            
                            for keyword in keywords:
                                try:
                                    cursor.execute(
                                        'INSERT INTO faq_list_keys (faq_id, keyword, keyword_type, weight) VALUES (%s, %s, %s, %s)',
                                        (faq_id, keyword, 'manual', 1.0)
                                    )
                                except Exception:
                                    # 关键词重复，跳过
                                    pass

                        success_count += 1

                    except Exception as e:
                        logger.error(f"❌ 导入第{row_idx}行失败: {str(e)}")
                        errors.append({'row': row_idx, 'reason': str(e)})
                        failed_count += 1
                        continue

                # 提交事务
                conn.commit()

        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

        # 重新加载知识库
        try:
            knowledge_service._load_knowledge_base()
            logger.info("✅ 知识库已重新加载")
        except Exception as e:
            logger.warning(f"⚠️ 知识库重新加载失败: {str(e)}")

        return jsonify({
            'success': True,
            'message': '导入完成',
            'data': {
                'total': total,
                'success_count': success_count,
                'failed_count': failed_count,
                'errors': errors
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ 导入FAQ失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/stats', methods=['GET'], endpoint='get_faq_stats')
@token_required
def get_faq_stats(user_id):
    """
    获取FAQ统计信息

    Headers:
        Authorization: Bearer <token>

    Returns:
        JSON: FAQ统计数据
    """
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # 总体统计
            cursor.execute('SELECT COUNT(*) as total FROM faq_list')
            total = cursor.fetchone()['total']

            # 状态统计
            cursor.execute('SELECT status, COUNT(*) as count FROM faq_list GROUP BY status')
            status_stats = cursor.fetchall()

            # 分类统计
            cursor.execute("""
                SELECT category, COUNT(*) as count
                FROM faq_list
                WHERE category IS NOT NULL AND category != ''
                GROUP BY category
                ORDER BY count DESC
                LIMIT 10
            """)
            category_stats = cursor.fetchall()

            # 来源统计
            cursor.execute("""
                SELECT source, COUNT(*) as count
                FROM faq_list
                WHERE source IS NOT NULL AND source != ''
                GROUP BY source
                ORDER BY count DESC
                LIMIT 10
            """)
            source_stats = cursor.fetchall()

            # 关键词统计
            cursor.execute('SELECT COUNT(*) as total_keywords FROM faq_list_keys')
            total_keywords = cursor.fetchone()['total_keywords']

        conn.close()

        return jsonify({
            'success': True,
            'data': {
                'total_faqs': total,
                'total_keywords': total_keywords,
                'status_distribution': status_stats,
                'top_categories': category_stats,
                'top_sources': source_stats
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ 获取FAQ统计失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@faq_bp.route('/import-template', methods=['GET'], endpoint='download_faq_import_template')
@token_required
def download_faq_import_template(user_id):
    """
    下载FAQ导入模板

    Headers:
        Authorization: Bearer <token>

    Returns:
        Excel文件: FAQ导入模板
    """
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "FAQ导入模板"

        # 设置表头
        headers = [
            'question',      # 问题（必填）
            'answer',        # 答案（必填）
            'category',      # 分类（可选）
            'source',        # 来源（可选）
            'keywords',      # 关键词（可选，逗号分隔）
            'status',        # 状态（可选，1=启用 0=禁用，默认1）
            'sort_order'     # 排序（可选，数字越小越靠前，默认0）
        ]
        
        ws.append(headers)

        # 添加示例数据
        example_rows = [
            [
                '什么是糖尿病？',
                '糖尿病是一种慢性代谢性疾病，主要特征是血糖水平持续升高。',
                '基础知识',
                '医学百科',
                '糖尿病,血糖,慢性病',
                1,
                1
            ],
            [
                '如何控制血糖？',
                '控制血糖需要：1.合理饮食 2.适量运动 3.按时服药 4.定期监测',
                '健康管理',
                '医生建议',
                '血糖控制,饮食,运动,用药',
                1,
                2
            ],
            [
                '糖尿病患者可以吃水果吗？',
                '可以适量吃水果，建议选择低糖水果如苹果、梨、柚子等，每天不超过200克。',
                '饮食指导',
                '营养师建议',
                '水果,饮食,血糖',
                1,
                3
            ]
        ]

        for row in example_rows:
            ws.append(row)

        # 设置列宽
        column_widths = {
            'A': 30,  # question
            'B': 50,  # answer
            'C': 15,  # category
            'D': 15,  # source
            'E': 30,  # keywords
            'F': 10,  # status
            'G': 10   # sort_order
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 添加说明工作表
        ws_info = wb.create_sheet("使用说明")
        instructions = [
            ['FAQ导入模板使用说明'],
            [''],
            ['1. 必填字段：'],
            ['   - question: 问题内容'],
            ['   - answer: 答案内容'],
            [''],
            ['2. 可选字段：'],
            ['   - category: 分类（如：基础知识、健康管理、饮食指导等）'],
            ['   - source: 来源（如：医学百科、医生建议、营养师建议等）'],
            ['   - keywords: 关键词，多个关键词用英文逗号分隔'],
            ['   - status: 状态，1表示启用，0表示禁用（默认为1）'],
            ['   - sort_order: 排序值，数字越小越靠前（默认为0）'],
            [''],
            ['3. 注意事项：'],
            ['   - 请勿修改表头名称'],
            ['   - 问题和答案不能为空'],
            ['   - 关键词之间用英文逗号分隔，不要有空格'],
            ['   - 删除示例数据后再填写您的数据'],
            ['   - 保存为.xlsx格式'],
            [''],
            ['4. 导入步骤：'],
            ['   - 填写完数据后保存文件'],
            ['   - 在FAQ管理页面点击"导入"按钮'],
            ['   - 选择填写好的Excel文件'],
            ['   - 等待导入完成'],
        ]

        for row in instructions:
            ws_info.append(row)

        # 设置说明页样式
        ws_info.column_dimensions['A'].width = 80
        title_font = Font(bold=True, size=14, color="4472C4")
        ws_info['A1'].font = title_font

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d')
        filename = f'faq_import_template_{timestamp}.xlsx'

        logger.info(f"✅ FAQ导入模板生成成功: {filename}")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"❌ 生成FAQ导入模板失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': f'生成模板失败: {str(e)}'
        }), 500
