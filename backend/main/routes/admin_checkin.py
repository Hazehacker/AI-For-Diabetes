"""
管理员打卡记录路由
~~~~~~~~~~~~~~~~

管理员视角的打卡记录管理API端点：
- 查询所有用户打卡记录（支持筛选）
- 导出打卡记录为Excel

作者: 智糖团队
日期: 2025-01-15
"""

from flask import request, jsonify, Blueprint, send_file
from utils.jwt_helper import no_auth_required as token_required
from utils.database import get_db_connection
from utils.logger import get_logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import io

logger = get_logger(__name__)

# 创建Blueprint
admin_checkin_bp = Blueprint('admin_checkin', __name__, url_prefix='/api/admin/checkin')


@admin_checkin_bp.route('/records', methods=['GET'], endpoint='get_admin_checkin_records')
@token_required
def get_admin_checkin_records(user_id):
    """
    获取所有用户的打卡记录（管理员接口）
    
    Headers:
        Authorization: Bearer <token>
    
    Query Parameters:
        start_date (optional): 开始日期，格式YYYY-MM-DD
        end_date (optional): 结束日期，格式YYYY-MM-DD
        user_id (optional): 用户ID筛选
        page (optional): 页码，默认1
        page_size (optional): 每页数量，默认20
    
    Returns:
        JSON: {
            "success": true,
            "data": {
                "total": 100,
                "page": 1,
                "page_size": 20,
                "records": [...]
            }
        }
    """
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        filter_user_id = request.args.get('user_id')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        
        # 验证分页参数
        if page < 1:
            page = 1
        if page_size < 1 or page_size > 100:
            page_size = 20
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if start_date:
            where_conditions.append("cr.timestamp >= %s")
            params.append(f"{start_date} 00:00:00")
        
        if end_date:
            where_conditions.append("cr.timestamp <= %s")
            params.append(f"{end_date} 23:59:59")
        
        if filter_user_id:
            where_conditions.append("cr.user_id = %s")
            params.append(int(filter_user_id))
        
        where_clause = " AND " + " AND ".join(where_conditions) if where_conditions else ""
        
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 查询总数
            count_sql = f"""
                SELECT COUNT(*) as total
                FROM checkin_records cr
                WHERE 1=1{where_clause}
            """
            cursor.execute(count_sql, tuple(params))
            total = cursor.fetchone()['total']
            
            # 查询记录（带分页）
            offset = (page - 1) * page_size
            query_sql = f"""
                SELECT 
                    cr.record_id,
                    cr.user_id,
                    u.username,
                    cr.checkin_type,
                    cr.checkin_value,
                    cr.glucose_status,
                    cr.feeling_text,
                    cr.timestamp,
                    cr.is_completed
                FROM checkin_records cr
                LEFT JOIN users u ON cr.user_id = u.user_id
                WHERE 1=1{where_clause}
                ORDER BY cr.timestamp DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(query_sql, tuple(params + [page_size, offset]))
            records = cursor.fetchall()
            
            # 格式化记录
            formatted_records = []
            for record in records:
                formatted_records.append({
                    'record_id': record['record_id'],
                    'user_id': record['user_id'],
                    'username': record['username'] or 'Unknown',
                    'checkin_type': record['checkin_type'],
                    'checkin_value': record['checkin_value'],
                    'glucose_status': record['glucose_status'],
                    'feeling_text': record['feeling_text'],
                    'timestamp': record['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if record['timestamp'] else None,
                    'is_completed': bool(record['is_completed'])
                })
            
            logger.info(f"✅ 管理员查询打卡记录成功: 总数={total}, 页码={page}, 每页={page_size}")
            
            return jsonify({
                'success': True,
                'data': {
                    'total': total,
                    'page': page,
                    'page_size': page_size,
                    'records': formatted_records
                }
            }), 200
            
        finally:
            cursor.close()
            conn.close()
    
    except ValueError as e:
        logger.error(f"❌ 参数错误: {str(e)}")
        return jsonify({
            'success': False,
            'code': 400,
            'message': f'参数错误: {str(e)}',
            'data': {}
        }), 400
    
    except Exception as e:
        logger.error(f"❌ 查询打卡记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'code': 500,
            'message': str(e),
            'data': {}
        }), 500


@admin_checkin_bp.route('/export', methods=['GET'], endpoint='export_admin_checkin_records')
@token_required
def export_admin_checkin_records(user_id):
    """
    导出打卡记录为Excel文件（管理员接口）
    
    Headers:
        Authorization: Bearer <token>
    
    Query Parameters:
        start_date (optional): 开始日期，格式YYYY-MM-DD
        end_date (optional): 结束日期，格式YYYY-MM-DD
        user_id (optional): 用户ID筛选
    
    Returns:
        Excel文件下载
    """
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        filter_user_id = request.args.get('user_id')
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if start_date:
            where_conditions.append("cr.timestamp >= %s")
            params.append(f"{start_date} 00:00:00")
        
        if end_date:
            where_conditions.append("cr.timestamp <= %s")
            params.append(f"{end_date} 23:59:59")
        
        if filter_user_id:
            where_conditions.append("cr.user_id = %s")
            params.append(int(filter_user_id))
        
        where_clause = " AND " + " AND ".join(where_conditions) if where_conditions else ""
        
        # 获取数据库连接
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 查询所有记录（不分页）
            query_sql = f"""
                SELECT 
                    cr.record_id,
                    cr.user_id,
                    u.username,
                    cr.checkin_type,
                    cr.checkin_value,
                    cr.glucose_status,
                    cr.feeling_text,
                    cr.timestamp,
                    cr.is_completed
                FROM checkin_records cr
                LEFT JOIN users u ON cr.user_id = u.user_id
                WHERE 1=1{where_clause}
                ORDER BY cr.timestamp DESC
            """
            cursor.execute(query_sql, tuple(params))
            records = cursor.fetchall()
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "打卡记录"
            
            # 设置表头
            headers = [
                'record_id', 'user_id', 'username', 'checkin_type', 
                'checkin_value', 'glucose_status', 'feeling_text', 'timestamp'
            ]
            header_names = [
                '记录ID', '用户ID', '用户名', '打卡类型', 
                '打卡值', '血糖状态', '感受描述', '打卡时间'
            ]
            
            # 写入表头
            for col_idx, header_name in enumerate(header_names, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row_idx, record in enumerate(records, start=2):
                ws.cell(row=row_idx, column=1, value=record['record_id'])
                ws.cell(row=row_idx, column=2, value=record['user_id'])
                ws.cell(row=row_idx, column=3, value=record['username'] or 'Unknown')
                ws.cell(row=row_idx, column=4, value=record['checkin_type'])
                ws.cell(row=row_idx, column=5, value=record['checkin_value'])
                ws.cell(row=row_idx, column=6, value=record['glucose_status'])
                ws.cell(row=row_idx, column=7, value=record['feeling_text'])
                ws.cell(row=row_idx, column=8, value=record['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if record['timestamp'] else '')
            
            # 调整列宽
            column_widths = [12, 10, 15, 15, 20, 12, 30, 20]
            for col_idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 生成文件名
            filename = f"checkin_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            logger.info(f"✅ 导出打卡记录成功: {len(records)} 条记录")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        finally:
            cursor.close()
            conn.close()
    
    except ValueError as e:
        logger.error(f"❌ 参数错误: {str(e)}")
        return jsonify({
            'success': False,
            'code': 400,
            'message': f'参数错误: {str(e)}',
            'data': {}
        }), 400
    
    except Exception as e:
        logger.error(f"❌ 导出打卡记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'code': 500,
            'message': str(e),
            'data': {}
        }), 500
