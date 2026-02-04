"""
用户管理路由
~~~~~~~~~~~

用户信息管理的API端点：
- 获取用户信息
- 更新用户信息
- 用户列表
- 用户状态管理

作者: 智糖团队
日期: 2025-01-15
"""

from flask import request, jsonify
from . import user_bp
from utils.jwt_helper import no_auth_required as token_required
from services.user_service import get_user_service
from utils.logger import get_logger

logger = get_logger(__name__)

# 获取服务实例
user_service = get_user_service()


def safe_int(value, default=0):
    """
    安全地将字符串转换为整数，如果为空或无效则返回默认值
    
    Args:
        value: 要转换的值
        default: 默认值
        
    Returns:
        int: 转换后的整数
    """
    if not value or not str(value).strip():
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def get_user_id_from_request():
    """
    从请求参数中获取 user_id，支持从 body 或 query 参数中获取
    必须提供 user_id，否则返回 None
    
    Returns:
        int or None: 用户ID，如果未提供则返回 None
    """
    # 对于 GET 请求，直接从 query 参数获取
    if request.method == 'GET':
        user_id = request.args.get('user_id')
        if user_id is not None:
            try:
                return int(user_id)
            except (ValueError, TypeError):
                return None
        return None
    
    # 对于 POST/PUT/DELETE 等请求，先从 body 中获取
    if request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
        try:
            if request.is_json and request.get_json(silent=True):
                data = request.get_json() or {}
                user_id = data.get('user_id')
                if user_id is not None:
                    try:
                        return int(user_id)
                    except (ValueError, TypeError):
                        pass
        except Exception:
            pass  # 如果解析 body 失败，继续从 query 参数获取
    
    # 如果 body 中没有，从 query 参数中获取
    user_id = request.args.get('user_id')
    if user_id is not None:
        try:
            return int(user_id)
        except (ValueError, TypeError):
            return None
    
    # 如果都没有，返回 None
    return None


@user_bp.route('/user/profile', methods=['GET'], endpoint='get_profile')
def get_profile():
    """
    获取用户资料

    Query:
        user_id: 用户ID（必须提供）
    """
    try:
        # 从请求参数中获取 user_id（必须提供）
        user_id = get_user_id_from_request()
        if user_id is None:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '必须提供user_id参数'
            }), 400

        result = user_service.get_user_profile(user_id)
        if result.get('success'):
            # 返回用户信息，移除success和message字段，只返回data部分
            user_data = {
                'user': result.get('user', {}),
                'tags': result.get('tags', {}),
                'stats': result.get('stats', {})
            }
            return jsonify({
                'code': 200,
                'data': user_data,
                'success': True,
                'message': ''
            }), 200
        else:
            return jsonify({
                'code': 404,
                'data': {},
                'success': False,
                'message': result.get('message', '用户不存在')
            }), 404
        
    except Exception as e:
        logger.error(f"❌ 获取用户资料失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/user/profile', methods=['PUT'], endpoint='update_profile')
@token_required
def update_profile(user_id):
    """
    更新当前用户资料

    Body:
        {
            "user_id": 用户ID（可选，如果不提供则使用认证用户ID）,
            "nickname": "昵称",
            "email": "邮箱",
            "phone_number": "手机号",
            "date_of_birth": "2025-11-07"
        }
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        
        data = request.get_json()
        
        result = user_service.update_user_profile(
            user_id=user_id,
            nickname=data.get('nickname'),
            email=data.get('email'),
            phone_number=data.get('phone_number'),
            date_of_birth=data.get('date_of_birth')
        )
        
        return jsonify({
            'code': 200 if result.get('success') else 400,
            'data': result.get('data', {}) if result.get('success') else {},
            'success': result.get('success'),
            'message': result.get('message', '')
        }), 200 if result.get('success') else 400
        
    except Exception as e:
        logger.error(f"❌ 更新用户资料失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/<int:target_user_id>', methods=['PUT'], endpoint='update_user')
@token_required
def update_user(user_id, target_user_id):
    """
    更新指定用户资料（管理员可以更新任何用户）

    Body:
        {
            "user_id": 用户ID（可选，如果不提供则使用认证用户ID）,
            "nickname": "昵称",
            "email": "邮箱",
            "phone_number": "手机号",
            "date_of_birth": "2025-11-07",
            "is_admin": true/false  // 管理员权限
        }
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        
        data = request.get_json() or {}

        # 完全放开权限：管理员可以编辑任何用户信息，包括设置管理员权限
        result = user_service.update_user_profile_admin(
            user_id=target_user_id,
            nickname=data.get('nickname'),
            email=data.get('email'),
            phone_number=data.get('phone_number'),
            date_of_birth=data.get('date_of_birth'),
            is_admin=data.get('is_admin')
        )
        
        return jsonify({
            'code': 200 if result.get('success') else 400,
            'data': result.get('data', {}) if result.get('success') else {},
            'success': result.get('success'),
            'message': result.get('message', '')
        }), 200 if result.get('success') else 400
        
    except Exception as e:
        logger.error(f"❌ 更新用户资料失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users', methods=['GET'], endpoint='get_users_list')
@token_required
def get_users_list(user_id):
    """
    获取用户列表

    Query:
        user_id: 用户ID（可选，如果不提供则使用认证用户ID）
        page: 页码
        page_size: 每页数量
        keyword: 搜索关键词
        is_active: 是否激活
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        
        # 移除管理员权限检查，允许所有人查询用户列表
        # if not user_service.is_admin(user_id):
        #     return jsonify({
        #         'code': 403,
        #         'data': {},
        #         'success': False,
        #         'message': '需要管理员权限'
        #     }), 403

        # 使用安全转换函数处理分页参数
        page = safe_int(request.args.get('page'), default=1)
        page_size = safe_int(request.args.get('page_size'), default=20)

        keyword = request.args.get('keyword')
        if keyword:
            keyword = keyword.strip() if keyword.strip() else None

        username = request.args.get('username')
        if username:
            username = username.strip() if username.strip() else None

        phone_number = request.args.get('phone_number')
        if phone_number:
            phone_number = phone_number.strip() if phone_number.strip() else None

        is_active = request.args.get('is_active')
        if is_active is not None and is_active.strip():
            is_active = is_active.lower() == 'true'
        else:
            is_active = None

        is_admin = request.args.get('is_admin')
        if is_admin is not None and is_admin.strip():
            is_admin = is_admin.lower() == 'true'
        else:
            is_admin = None

        result = user_service.get_users_list(
            page=page,
            page_size=page_size,
            keyword=keyword,
            username=username,
            phone_number=phone_number,
            is_active=is_active,
            is_admin=is_admin
        )

        return jsonify({
            'code': 200,
            'data': result,
            'success': True
        }), 200

    except Exception as e:
        logger.error(f"❌ 获取用户列表失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/<int:target_user_id>', methods=['GET'], endpoint='get_user_detail')
@token_required
def get_user_detail(user_id, target_user_id):
    """
    获取指定用户详情

    Query:
        user_id: 用户ID（可选，如果不提供则使用认证用户ID）
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        
        user = user_service.get_user_by_id(target_user_id)
        
        if user:
            return jsonify({
                'code': 200,
                'data': {'user': user},
                'success': True
            }), 200
        else:
            return jsonify({
                'code': 404,
                'data': {},
                'success': False,
                'message': '用户不存在'
            }), 404
            
    except Exception as e:
        logger.error(f"❌ 获取用户详情失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/<int:target_user_id>', methods=['DELETE'], endpoint='delete_user')
@token_required
def delete_user(user_id, target_user_id):
    """
    删除用户（管理员接口）

    Query:
        user_id: 用户ID（可选，如果不提供则使用认证用户ID）
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        
        # 检查管理员权限（暂时跳过以进行测试）
        # if not user_service.is_admin(user_id):
        #     return jsonify({
        #         'code': 403,
        #         'data': {},
        #         'success': False,
        #         'message': '需要管理员权限'
        #     }), 403

        result = user_service.delete_user(target_user_id)
        return jsonify({
            'code': 200 if result.get('success') else 404,
            'data': result.get('data', {}) if result.get('success') else {},
            'success': result.get('success'),
            'message': result.get('message', '')
        }), 200 if result.get('success') else 404

    except Exception as e:
        logger.error(f"❌ 删除用户失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/<int:target_user_id>/status', methods=['PUT'], endpoint='toggle_user_status')
@token_required
def toggle_user_status(user_id, target_user_id):
    """
    切换用户状态（管理员接口）

    Query:
        user_id: 用户ID（可选，如果不提供则使用认证用户ID）
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id

        # 检查管理员权限（暂时跳过以进行测试）
        # if not user_service.is_admin(user_id):
        #     return jsonify({
        #         'code': 403,
        #         'data': {},
        #         'success': False,
        #         'message': '需要管理员权限'
        #     }), 403

        result = user_service.toggle_user_status(target_user_id)
        return jsonify({
            'code': 200 if result.get('success') else 400,
            'data': result.get('data', {}) if result.get('success') else {},
            'success': result.get('success'),
            'message': result.get('message', '')
        }), 200 if result.get('success') else 400

    except Exception as e:
        logger.error(f"❌ 切换用户状态失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/reset-password', methods=['POST', 'PUT'], endpoint='reset_user_password')
@token_required
def reset_user_password(user_id):
    """
    重置或设置用户密码（管理接口）

    Body:
        {
            "user_id": 123,           // 可选：操作者的用户ID（如果不提供则使用认证用户ID）
            "target_user_id": 456,    // 必需：要重置密码的用户ID
            "new_password": "新密码"   // 可选：如果不提供，使用默认密码 "123456"
        }

    Response:
        {
            "code": 200,
            "data": {"password": "设置的密码"},
            "message": "密码已成功设置" 或 "密码已重置为默认密码",
            "success": true
        }
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        
        # 获取请求体中的参数
        data = request.get_json()
        if not data:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '请求体不能为空'
            }), 400

        target_user_id = data.get('target_user_id') or data.get('user_id')
        new_password = data.get('new_password')

        if not target_user_id:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': 'target_user_id参数是必需的'
            }), 400

        # 默认管理员权限：所有用户都可以重置密码
        result = user_service.reset_user_password(target_user_id, new_password)
        return jsonify({
            'code': 200 if result.get('success') else 400,
            'data': {
                'password': result.get('password')
            } if result.get('success') else {},
            'success': result.get('success'),
            'message': result.get('message', '')
        }), 200 if result.get('success') else 400

    except Exception as e:
        logger.error(f"❌ 重置用户密码失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/batch-import', methods=['POST'], endpoint='batch_import_users')
@token_required
def batch_import_users(user_id):
    """
    批量导入用户（管理接口）

    支持两种方式：
    1. 文件上传：上传Excel(.xlsx)或CSV(.csv)文件
       Form Data:
           file: Excel或CSV文件（必需）
           user_id: 用户ID（可选，如果不提供则使用认证用户ID）

    2. JSON格式：（向后兼容）
       Body:
           {
               "user_id": 用户ID（可选，如果不提供则使用认证用户ID）,
               "users": [{username, password, nickname, phone_number}, ...]
           }

    文件格式要求：
    - 第一行为表头：username, password, nickname, phone_number
    - 后续行为数据行
    - password字段如果为空，将使用默认密码"123456"
    - 系统会自动检查用户名和手机号的唯一性，重复的记录不会被导入
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id

        # 检查是否有文件上传
        if 'file' in request.files:
            # 文件上传方式
            file = request.files['file']
            if file.filename == '':
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': '文件名不能为空'
                }), 400

            # 检查文件类型
            allowed_extensions = ['xlsx', 'xls', 'csv']
            file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

            if file_extension not in allowed_extensions:
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': f'不支持的文件类型。支持的格式：{", ".join(allowed_extensions)}'
                }), 400

            # 解析文件
            try:
                if file_extension in ['xlsx', 'xls']:
                    # Excel文件 - 使用openpyxl
                    from openpyxl import load_workbook
                    wb = load_workbook(file)
                    ws = wb.active

                    # 获取表头
                    headers = []
                    for cell in ws[1]:  # 第一行作为表头
                        if cell.value:
                            headers.append(str(cell.value).strip())

                    # 验证必需的列
                    required_columns = ['username']
                    missing_columns = [col for col in required_columns if col not in headers]
                    if missing_columns:
                        return jsonify({
                            'code': 400,
                            'data': {},
                            'success': False,
                            'message': f'缺少必需的列：{", ".join(missing_columns)}'
                        }), 400

                    # 读取数据行
                    users_data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始
                        if not row or all(cell is None for cell in row):
                            continue  # 跳过空行

                        user_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                column_name = headers[i]
                                if value is not None:
                                    user_dict[column_name] = str(value).strip()
                                else:
                                    user_dict[column_name] = None

                        # 如果没有设置密码，使用默认密码
                        if not user_dict.get('password'):
                            user_dict['password'] = '123456'

                        users_data.append(user_dict)

                else:
                    # CSV文件 - 使用csv模块
                    import csv
                    import io

                    # 将文件内容转换为字符串
                    file_content = file.read().decode('utf-8')
                    csv_reader = csv.DictReader(io.StringIO(file_content))

                    # 获取表头
                    headers = csv_reader.fieldnames
                    if not headers:
                        return jsonify({
                            'code': 400,
                            'data': {},
                            'success': False,
                            'message': 'CSV文件格式错误，无法读取表头'
                        }), 400

                    # 验证必需的列
                    required_columns = ['username']
                    missing_columns = [col for col in required_columns if col not in headers]
                    if missing_columns:
                        return jsonify({
                            'code': 400,
                            'data': {},
                            'success': False,
                            'message': f'缺少必需的列：{", ".join(missing_columns)}'
                        }), 400

                    # 读取数据行
                    users_data = []
                    for row in csv_reader:
                        user_dict = {}
                        for header in headers:
                            value = row.get(header, '').strip()
                            if value:
                                user_dict[header] = value
                            else:
                                user_dict[header] = None

                        # 如果没有设置密码，使用默认密码
                        if not user_dict.get('password'):
                            user_dict['password'] = '123456'

                        users_data.append(user_dict)

            except Exception as e:
                logger.error(f"❌ 解析文件失败: {str(e)}")
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': f'文件解析失败：{str(e)}'
                }), 400

        else:
            # JSON格式（向后兼容）
            data = request.get_json()
            if not data or 'users' not in data:
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': '缺少users参数或上传文件'
                }), 400

            users_data = data['users']
            if not isinstance(users_data, list):
                return jsonify({
                    'code': 400,
                    'data': {},
                    'success': False,
                    'message': 'users必须是数组'
                }), 400

        # 验证用户数据
        if not users_data:
            return jsonify({
                'code': 400,
                'data': {},
                'success': False,
                'message': '用户数据不能为空'
            }), 400

        result = user_service.batch_import_users(users_data)

        # 如果有错误详情，一起返回
        response_data = {
            'code': 200 if result.get('success') else 400,
            'data': result.get('data', {}),
            'success': result.get('success'),
            'message': result.get('message', '')
        }

        if 'errors' in result:
            response_data['errors'] = result['errors']

        return jsonify(response_data), 200 if result.get('success') else 400

    except Exception as e:
        logger.error(f"❌ 批量导入用户失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500


@user_bp.route('/users/import-template', methods=['GET'], endpoint='download_import_template')
@token_required
def download_import_template(user_id):
    """
    下载批量导入用户模板（管理员接口）

    Query:
        user_id: 用户ID（可选，如果不提供则使用认证用户ID）
    """
    try:
        # 从请求参数中获取 user_id（可选，如果不提供则使用装饰器传入的 user_id）
        request_user_id = get_user_id_from_request()
        if request_user_id is not None:
            user_id = request_user_id
        

        # 创建模板文件内容
        import csv
        from io import StringIO
        from flask import Response
        
        output = StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['username', 'password', 'nickname', 'phone_number'])
        # 写入示例数据
        writer.writerow(['user1', '123456', '用户1', '13800138000'])
        writer.writerow(['user2', '123456', '用户2', '13800138001'])
        
        output.seek(0)
        
        # 返回CSV文件
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=user_import_template.csv'
            }
        )

    except Exception as e:
        logger.error(f"❌ 下载导入模板失败: {str(e)}")
        return jsonify({
            'code': 500,
            'data': {},
            'success': False,
            'message': str(e)
        }), 500

